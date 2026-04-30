"""
数据解析流水线：上传校验 → 数据清洗 → 通用表格解析（任意格式 Excel）→ 字段画像与聚合。
不限制行列格式，支持多 sheet；产出 tables（表头+行数据）、table_schemas（供 read_metrics/LLM）、
column_profiles（每列画像）、aggregations（时间序列与按维度聚合），并为后续自动/按需看板生成提供基础。
"""
import io
import logging
from typing import Any, Dict, List, Tuple

logger = logging.getLogger(__name__)

MAX_SHEETS = 20
MAX_ROWS_PER_SHEET = 5000
MAX_COLS = 200


def _cell(row: tuple, i: int) -> str:
    if row is None or i >= len(row):
        return ""
    v = row[i]
    return "" if v is None else str(v).strip()


def _cell_any(row: tuple, i: int) -> Any:
    if row is None or i >= len(row):
        return None
    return row[i]


def validate_upload(content: bytes, filename: str) -> None:
    """上传校验：文件类型、可读性。"""
    if not filename or not filename.lower().endswith((".xlsx", ".xls")):
        raise ValueError("请上传 .xlsx 或 .xls 文件")
    if not content or len(content) < 100:
        raise ValueError("文件过小或为空")
    try:
        import openpyxl
        openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True).close()
    except Exception as e:
        raise ValueError(f"无法解析 Excel：{e}") from e


def _clean_cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return v


def _clean_rows(rows: list[tuple], max_cols: int) -> list[list[Any]]:
    """去空行、统一列数、清洗单元格。"""
    cleaned = []
    for row in rows:
        filled = [_clean_cell(_cell_any(row, i)) for i in range(max_cols)]
        if all(v == "" or v is None for v in filled):
            continue
        cleaned.append(filled)
    return cleaned


def _to_str_list(row: list[Any]) -> list[str]:
    return [str(x) if x is not None and x != "" else "" for x in row]


def _detect_used_col_count(rows: list[tuple], default_cols: int) -> int:
    """
    从原始 rows 中检测“实际使用列数”（最右非空列索引 + 1），
    避免因 iter_rows(max_col=200) 导致所有 sheet 都被扩成 200 列。
    """
    used = 0
    for row in rows or []:
        if not row:
            continue
        last = -1
        for i, v in enumerate(row):
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            last = i
        if last + 1 > used:
            used = last + 1
    if used <= 0:
        return max(1, min(default_cols, MAX_COLS))
    return min(used, MAX_COLS)


def _normalize_headers(headers: list[str]) -> list[str]:
    """
    表头规范化：
    - 空列名补成“列N”
    - 重名列追加后缀“_2/_3”
    """
    out: list[str] = []
    seen: dict[str, int] = {}
    for i, h in enumerate(headers):
        base = (h or "").strip()
        if not base:
            base = f"列{i + 1}"
        n = int(seen.get(base) or 0) + 1
        seen[base] = n
        out.append(base if n == 1 else f"{base}_{n}")
    return out


def _looks_like_header_vs_data(header_row: list[Any], data_row: list[Any]) -> bool:
    """简单判断一行是否更像表头：自身多为字符串，下一行更多是数字。"""
    if not header_row or not data_row:
        return False
    header_strings = 0
    for v in header_row:
        if isinstance(v, str) and not v.strip().replace(".", "").replace("-", "").isdigit():
            header_strings += 1
    data_numbers = 0
    for v in data_row:
        if isinstance(v, (int, float)):
            data_numbers += 1
        elif isinstance(v, str) and v.strip().replace(".", "").replace("-", "").isdigit():
            data_numbers += 1
    return header_strings >= max(1, int(len(header_row) * 0.6)) and data_numbers >= max(1, int(len(data_row) * 0.3))


def _detect_header_row(cleaned_rows: list[list[Any]]) -> int | None:
    """
    在清洗后的前若干行中尝试识别表头行（返回在 cleaned_rows 中的索引，0-based）。
    参考 chatExcel-mcp 的 ExcelStructureDetector，但做了轻量化：
    - 有至少两个非空单元格；
    - 包含典型标题关键词；或
    - 与下一行在“字符串 vs 数字分布”上明显不同。
    """
    max_check = min(20, len(cleaned_rows))
    header_indicators = [
        "名称",
        "姓名",
        "name",
        "id",
        "编号",
        "日期",
        "date",
        "时间",
        "月份",
        "month",
        "项目",
        "部门",
        "金额",
        "amount",
        "收入",
        "营收",
        "利润",
        "成本",
        "费用",
    ]
    best_idx = None
    best_score = -1.0
    for i in range(max_check):
        row = cleaned_rows[i]
        non_empty = [str(v).strip() for v in row if v not in (None, "")]
        if len(non_empty) < 2:
            continue
        text = "".join(non_empty).lower()
        if any(ind in text for ind in header_indicators):
            return i
        if i + 1 < len(cleaned_rows):
            if _looks_like_header_vs_data(row, cleaned_rows[i + 1]):
                return i
        # 兜底打分：字符串占比高 + 唯一值比例高 + 下一行有更多数值
        str_cnt = 0
        num_next = 0
        for v in row:
            s = str(v).strip() if v is not None else ""
            if s and not s.replace(".", "").replace("-", "").isdigit():
                str_cnt += 1
        if i + 1 < len(cleaned_rows):
            for v in cleaned_rows[i + 1]:
                s = str(v).strip() if v is not None else ""
                if s and s.replace(".", "").replace("-", "").isdigit():
                    num_next += 1
        uniq_ratio = len(set(non_empty)) / max(1, len(non_empty))
        score = str_cnt * 1.0 + uniq_ratio * 2.0 + num_next * 0.3
        if score > best_score:
            best_score = score
            best_idx = i
    if best_idx is not None and best_score >= 2.5:
        return best_idx
    return None


def generic_parse_excel(content: bytes) -> dict[str, Any]:
    """
    通用解析：任意格式 Excel，不限制行列结构。
    每个 sheet 先进行轻量结构检测，自动推断表头行，再抽取数据行；多 sheet 均解析。
    返回 { "tables": { sheet_name: { "headers": [...], "rows": [[...], ...] } }, "table_schemas": [...] }。
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl 未安装")
        return {"tables": {}, "table_schemas": []}

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("解析 Excel 失败: %s", e)
        return {"tables": {}, "table_schemas": []}

    tables: dict[str, dict[str, Any]] = {}
    table_schemas: list[dict[str, Any]] = []
    sheet_names = getattr(wb, "sheetnames", []) or []
    main_sheet: str | None = None
    main_sheet_rows = -1

    for idx, name in enumerate(sheet_names):
        if idx >= MAX_SHEETS:
            break
        try:
            ws = wb[name]
        except Exception:
            continue
        all_rows = list(
            ws.iter_rows(min_row=1, max_row=MAX_ROWS_PER_SHEET + 1, max_col=MAX_COLS, values_only=True)
        )
        if not all_rows:
            tables[name] = {"headers": [], "rows": []}
            table_schemas.append({"sheet_name": name, "headers": [], "row_count": 0})
            continue
        max_cols_num = max(len(r) for r in all_rows) if all_rows else 0
        max_cols_num = min(max_cols_num, MAX_COLS)
        max_cols_num = _detect_used_col_count(all_rows, max_cols_num)
        cleaned = _clean_rows(all_rows, max_cols_num)
        if not cleaned:
            tables[name] = {"headers": [], "rows": []}
            table_schemas.append({"sheet_name": name, "headers": [], "row_count": 0})
            continue

        header_idx = _detect_header_row(cleaned)
        if header_idx is None:
            header_idx = 0
        headers = _normalize_headers(_to_str_list(cleaned[header_idx]))
        data_block = cleaned[header_idx + 1 :]
        rows = [list(row)[: len(headers)] for row in data_block]
        if len(rows) > MAX_ROWS_PER_SHEET:
            rows = rows[:MAX_ROWS_PER_SHEET]
        tables[name] = {"headers": headers, "rows": rows}
        row_count = len(rows)
        table_schemas.append({"sheet_name": name, "headers": headers, "row_count": row_count})
        if row_count > main_sheet_rows:
            main_sheet_rows = row_count
            main_sheet = name

    try:
        wb.close()
    except Exception:
        pass
    # 标记主表（数据行数最多的 sheet）
    if main_sheet is not None:
        for schema in table_schemas:
            if schema.get("sheet_name") == main_sheet:
                schema["is_main_sheet"] = True
            else:
                schema["is_main_sheet"] = False
    return {"tables": tables, "table_schemas": table_schemas}


def _build_validation_summary(tables: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """
    轻量级数据质量与完整性检查结果：
    - row_count / col_count
    - missing_ratio: 所有单元格中空值占比
    - duplicate_ratio: 重复行占比
    - mixed_type_columns: 同一列中存在多种 Python 类型的列名列表
    """
    summary: Dict[str, Dict[str, Any]] = {}
    for sheet, t in (tables or {}).items():
        headers: List[str] = t.get("headers") or []
        rows: List[List[Any]] = t.get("rows") or []
        row_count = len(rows)
        col_count = len(headers)
        total_cells = row_count * col_count
        missing = 0
        if total_cells > 0:
            for r in rows:
                for v in r[:col_count]:
                    if v is None or v == "":
                        missing += 1
        missing_ratio = float(missing) / float(total_cells) if total_cells > 0 else 0.0

        # 重复行占比（按整行值比较）
        duplicate_ratio = 0.0
        if row_count > 0:
            seen = set()
            dup_count = 0
            for r in rows:
                key = tuple(r[:col_count])
                if key in seen:
                    dup_count += 1
                else:
                    seen.add(key)
            duplicate_ratio = float(dup_count) / float(row_count)

        # 混合类型列
        mixed_type_columns: List[str] = []
        for idx, name in enumerate(headers):
            types: set[str] = set()
            for r in rows:
                if idx >= len(r):
                    continue
                v = r[idx]
                if v is None or v == "":
                    continue
                types.add(type(v).__name__)
                if len(types) > 1:
                    mixed_type_columns.append(name)
                    break

        summary[sheet] = {
            "row_count": row_count,
            "col_count": col_count,
            "missing_ratio": missing_ratio,
            "duplicate_ratio": duplicate_ratio,
            "mixed_type_columns": mixed_type_columns,
        }
    return summary


def _infer_number(v: Any) -> bool:
    if v is None or v == "":
        return False
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        try:
            float(v)
            return True
        except ValueError:
            return False
    return False


def _infer_column_profiles(tables: Dict[str, Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """
    对每个 sheet 的每一列做轻量画像：
    - type: number/string
    - null_ratio: 空值比例
    - distinct_count: 不同取值数
    - is_time: 是否疑似时间列（列名含日期/时间关键词，或值类似日期）
    - is_dimension: 是否维度列（离散值有限）
    - is_metric: 是否指标列（主要为数值）
    """
    profiles: Dict[str, List[Dict[str, Any]]] = {}
    max_sample_rows = 1000
    for sheet, t in (tables or {}).items():
        headers: List[str] = t.get("headers") or []
        rows: List[List[Any]] = t.get("rows") or []
        if not headers:
            profiles[sheet] = []
            continue
        col_count = len(headers)
        col_values: List[List[Any]] = [[] for _ in range(col_count)]
        non_null_counts = [0] * col_count
        number_counts = [0] * col_count
        total_rows = min(len(rows), max_sample_rows)
        for r in rows[:total_rows]:
            for idx in range(col_count):
                v = r[idx] if idx < len(r) else None
                if v is None or v == "":
                    continue
                non_null_counts[idx] += 1
                col_values[idx].append(v)
                if _infer_number(v):
                    number_counts[idx] += 1
        sheet_profiles: List[Dict[str, Any]] = []
        for idx, name in enumerate(headers):
            nn = non_null_counts[idx]
            nums = number_counts[idx]
            distinct_count = len({str(v) for v in col_values[idx]}) if col_values[idx] else 0
            null_ratio = 0.0
            if total_rows > 0:
                null_ratio = 1.0 - nn / float(total_rows)
            is_number_col = nn > 0 and nums / float(nn) >= 0.8
            lower_name = (name or "").lower()
            is_time_name = any(key in lower_name for key in ["date", "day", "month", "year", "日期", "时间"])
            # 这里只做列名启发式，不根据具体值判断日期，避免文化/格式差异
            is_time = bool(is_time_name)
            # 维度列：非数值，或 distinct_count 较小
            is_dimension = (not is_number_col and distinct_count > 0 and distinct_count <= min(50, max(1, total_rows // 2)))
            # 指标列：主要为数值，且 distinct_count>1
            is_metric = bool(is_number_col and distinct_count > 1)
            sheet_profiles.append(
                {
                    "name": name,
                    "type": "number" if is_number_col else "string",
                    "null_ratio": null_ratio,
                    "distinct_count": distinct_count,
                    "is_time": is_time,
                    "is_dimension": is_dimension,
                    "is_metric": is_metric,
                }
            )
        profiles[sheet] = sheet_profiles
    return profiles


def _build_aggregations(
    tables: Dict[str, Dict[str, Any]], profiles: Dict[str, List[Dict[str, Any]]]
) -> Dict[str, Dict[str, Any]]:
    """
    基于字段画像构建轻量聚合结果：
    - time_series: 若存在时间列 + 指标列，按时间分组求和
    - by_dimension: 若存在维度列 + 指标列，按维度分组求和（取前 N 个）
    """
    aggs: Dict[str, Dict[str, Any]] = {}
    top_n = 10
    for sheet, t in (tables or {}).items():
        headers: List[str] = t.get("headers") or []
        rows: List[List[Any]] = t.get("rows") or []
        profs = profiles.get(sheet) or []
        if not headers or not rows or not profs:
            aggs[sheet] = {}
            continue
        time_cols = [p["name"] for p in profs if p.get("is_time")]
        metric_cols = [p["name"] for p in profs if p.get("is_metric")]
        dim_cols = [p["name"] for p in profs if p.get("is_dimension")]
        sheet_aggs: Dict[str, Any] = {}
        # 时间序列：选第一个时间列 + 若干指标列
        if time_cols and metric_cols:
            t_col = time_cols[0]
            ts_map: Dict[str, Dict[str, float]] = {}
            t_idx = headers.index(t_col) if t_col in headers else -1
            m_indices: List[Tuple[str, int]] = []
            for m_name in metric_cols[:3]:
                if m_name in headers:
                    m_indices.append((m_name, headers.index(m_name)))
            if t_idx >= 0 and m_indices:
                for r in rows:
                    if t_idx >= len(r):
                        continue
                    t_key = str(r[t_idx]) if r[t_idx] is not None and r[t_idx] != "" else ""
                    if t_key == "":
                        continue
                    ts_entry = ts_map.setdefault(t_key, {name: 0.0 for name, _ in m_indices})
                    for m_name, mi in m_indices:
                        if mi < len(r):
                            from .data_parse_chat import _to_num  # type: ignore

                            ts_entry[m_name] += _to_num(r[mi])
                x_labels = sorted(ts_map.keys())
                series = []
                for m_name, _ in m_indices:
                    series.append({"name": m_name, "data": [ts_map[x].get(m_name, 0.0) for x in x_labels]})
                if x_labels and series:
                    sheet_aggs["time_series"] = {"time_column": t_col, "labels": x_labels, "series": series}
        # 维度聚合：每个维度 + 若干指标列
        by_dim_list: List[Dict[str, Any]] = []
        if dim_cols and metric_cols:
            for d_col in dim_cols[:3]:
                d_idx = headers.index(d_col) if d_col in headers else -1
                if d_idx < 0:
                    continue
                m_indices: List[Tuple[str, int]] = []
                for m_name in metric_cols[:3]:
                    if m_name in headers:
                        m_indices.append((m_name, headers.index(m_name)))
                if not m_indices:
                    continue
                dim_map: Dict[str, Dict[str, float]] = {}
                for r in rows:
                    if d_idx >= len(r):
                        continue
                    key = str(r[d_idx]) if r[d_idx] is not None and r[d_idx] != "" else ""
                    if key == "":
                        continue
                    agg_entry = dim_map.setdefault(key, {name: 0.0 for name, _ in m_indices})
                    for m_name, mi in m_indices:
                        if mi < len(r):
                            from .data_parse_chat import _to_num  # type: ignore

                            agg_entry[m_name] += _to_num(r[mi])
                # 取前 top_n 维度（按第一个指标排序）
                if not dim_map:
                    continue
                first_metric = m_indices[0][0]
                sorted_items = sorted(dim_map.items(), key=lambda kv: kv[1].get(first_metric, 0.0), reverse=True)[
                    :top_n
                ]
                labels = [k for k, _ in sorted_items]
                series = []
                for m_name, _ in m_indices:
                    series.append({"name": m_name, "data": [dim_map[k].get(m_name, 0.0) for k in labels]})
                if labels and series:
                    by_dim_list.append({"dimension": d_col, "labels": labels, "series": series})
        if by_dim_list:
            sheet_aggs["by_dimension"] = by_dim_list
        aggs[sheet] = sheet_aggs
    return aggs


def _build_auto_dashboards(aggregations: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    根据聚合结果构建默认看板图表（auto_dashboards），供前端直接用 Recharts 渲染。
    """
    dashboards: List[Dict[str, Any]] = []
    for sheet, agg in (aggregations or {}).items():
        ts = agg.get("time_series")
        if ts:
            labels = ts.get("labels") or []
            option = {
                "xAxis": {"type": "category", "data": labels},
                "yAxis": {"type": "value"},
                "series": [
                    {"name": s.get("name"), "type": "line", "data": s.get("data") or []}
                    for s in ts.get("series") or []
                ],
            }
            dashboards.append(
                {
                    "id": f"{sheet}_time_series",
                    "title": f"{sheet}: {ts.get('time_column', '')} 趋势",
                    "option": option,
                }
            )
        for by_dim in agg.get("by_dimension") or []:
            labels = by_dim.get("labels") or []
            option = {
                "xAxis": {"type": "category", "data": labels},
                "yAxis": {"type": "value"},
                "series": [
                    {"name": s.get("name"), "type": "bar", "data": s.get("data") or []}
                    for s in by_dim.get("series") or []
                ],
            }
            dashboards.append(
                {
                    "id": f"{sheet}_by_{by_dim.get('dimension', '')}",
                    "title": f"{sheet}: 按 {by_dim.get('dimension', '')} 聚合",
                    "option": option,
                }
            )
    return dashboards


def run_pipeline(content: bytes, filename: str) -> dict[str, Any]:
    """
    执行完整流水线：校验 → 通用解析 → 字段画像与聚合 → 默认看板。
    返回 {
      "tables": {...},
      "table_schemas": [...],
      "column_profiles": {...},
      "aggregations": {...},
      "auto_dashboards": [...],
      "kanban_config": [...],
    }。
    """
    validate_upload(content, filename)
    parsed = generic_parse_excel(content)
    tables = parsed.get("tables") or {}
    table_schemas = parsed.get("table_schemas") or []
    column_profiles = _infer_column_profiles(tables)
    aggregations = _build_aggregations(tables, column_profiles)
    auto_dashboards = _build_auto_dashboards(aggregations)
    validation_summary = _build_validation_summary(tables)
    return {
        "tables": tables,
        "table_schemas": table_schemas,
        "column_profiles": column_profiles,
        "aggregations": aggregations,
        "auto_dashboards": auto_dashboards,
        "validation_summary": validation_summary,
        # 向后兼容：kanban_config 默认等于 auto_dashboards
        "kanban_config": auto_dashboards,
    }
