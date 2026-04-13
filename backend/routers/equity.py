import csv
import io
import json
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass
from io import BytesIO
import zipfile
from datetime import datetime
from typing import Any
from pathlib import Path

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import bindparam, text

from backend.database import get_db
from backend.geo_cn import city_display_label, normalize_geo, province_display_label

router = APIRouter()

DEBUG_LOG_PATH = Path("debug-6b39de.log")


def _dbg(hypothesis_id: str, location: str, message: str, data: dict):
    # NDJSON append; avoid secrets/PII
    try:
        import json as _json, time as _time

        payload = {
            "sessionId": "6b39de",
            "runId": "pre-fix",
            "hypothesisId": hypothesis_id,
            "location": location,
            "message": message,
            "data": data,
            "timestamp": int(_time.time() * 1000),
        }
        DEBUG_LOG_PATH.open("a", encoding="utf-8").write(_json.dumps(payload, ensure_ascii=False) + "\n")
    except Exception:
        pass


def _now_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def _csv_get(row: dict[str, Any], key: str) -> str:
    v = row.get(key)
    return str(v).strip() if v is not None else ""


def _upsert_equity_target(
    db: Any,
    *,
    snapshot_name: str,
    entity_id: str,
    name: str,
    credit_code: Any,
    alias: Any,
    is_key: bool,
    notes: Any,
) -> None:
    """
    同一 snapshot 下每个 entity_id 只保留一条 target：已存在则合并（is_key 取或、名称取更长者）。
    """
    row = db.execute(
        text(
            """
            SELECT id, is_key, name
            FROM equity_targets
            WHERE snapshot_name = :s AND entity_id = :eid
            LIMIT 1
            """
        ),
        {"s": snapshot_name, "eid": entity_id},
    ).fetchone()
    if row:
        tid = str(row[0])
        old_key = bool(row[1])
        old_name = str(row[2] or "")
        merged_key = old_key or is_key
        merged_name = name.strip() if len(name.strip()) >= len(old_name.strip()) else old_name
        db.execute(
            text(
                """
                UPDATE equity_targets
                SET name = :name,
                    is_key = :ik,
                    credit_code = COALESCE(:cc, credit_code),
                    alias = COALESCE(:alias, alias),
                    notes = COALESCE(:notes, notes)
                WHERE id = :tid
                """
            ),
            {
                "tid": tid,
                "name": merged_name,
                "ik": merged_key,
                "cc": credit_code,
                "alias": alias,
                "notes": notes,
            },
        )
        return
    db.execute(
        text(
            """
            INSERT INTO equity_targets
              (id, snapshot_name, entity_id, name, credit_code, alias, is_key, notes)
            VALUES
              (:id, :s, :eid, :name, :cc, :alias, :key, :notes)
            """
        ),
        {
            "id": str(uuid.uuid4()),
            "s": snapshot_name,
            "eid": entity_id,
            "name": name,
            "cc": credit_code,
            "alias": alias,
            "key": is_key,
            "notes": notes,
        },
    )


def _read_zip_member_text(zf: zipfile.ZipFile, member_name: str) -> str | None:
    """
    从 zip 中读取指定成员（按 basename 匹配），返回 UTF-8 文本。
    """
    target = member_name.strip().lower()
    for info in zf.infolist():
        if info.is_dir():
            continue
        base = (info.filename.split("/")[-1].split("\\")[-1] or "").strip().lower()
        if base == target:
            raw = zf.read(info)
            try:
                return raw.decode("utf-8-sig")
            except Exception:
                return raw.decode("utf-8", errors="replace")
    return None


def _parse_float(val: Any) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    # 兼容 "20%" / "20.5 %"
    s = s.replace("%", "").strip()
    try:
        return float(s)
    except Exception:
        return None


def _pick_col(headers: list[str], keywords: list[str]) -> int | None:
    for i, h in enumerate(headers):
        hh = (h or "").strip()
        if not hh:
            continue
        for kw in keywords:
            if kw in hh:
                return i
    return None


@dataclass
class _TyEdge:
    from_name: str
    to_name: str
    hold_pct: float | None
    hold_pct_text: str | None


def _read_tianyancha_xlsx_edges(raw: bytes, company_name: str, kind: str) -> list[_TyEdge]:
    """
    天眼查导出（xlsx）轻量解析：
    - kind=shareholders：股东 -> 公司
    - kind=investments：公司 -> 被投
    """
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"openpyxl not available: {e}")

    try:
        wb = load_workbook(BytesIO(raw), data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"invalid xlsx: {e}")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    headers = [str(x).strip() if x is not None else "" for x in header_row]
    if kind == "shareholders":
        name_col = _pick_col(headers, ["股东", "股东名称", "出资人", "投资方"])
        pct_col = _pick_col(headers, ["持股比例", "持股", "占比", "出资比例"])
        if name_col is None:
            raise HTTPException(status_code=400, detail="无法识别股东名称列（请确认导出为股东信息表）")
        out: list[_TyEdge] = []
        for r in rows_iter:
            if not r:
                continue
            n = str(r[name_col]).strip() if name_col < len(r) and r[name_col] is not None else ""
            if not n:
                continue
            pct_val = r[pct_col] if (pct_col is not None and pct_col < len(r)) else None
            pct = _parse_float(pct_val)
            pct_text = str(pct_val).strip() if pct_val is not None else None
            out.append(_TyEdge(from_name=n, to_name=company_name, hold_pct=pct, hold_pct_text=pct_text))
        return out

    if kind == "investments":
        to_col = _pick_col(headers, ["被投资企业", "被投企业", "投资企业", "企业名称", "被投资方"])
        pct_col = _pick_col(headers, ["投资比例", "持股比例", "占比"])
        if to_col is None:
            raise HTTPException(status_code=400, detail="无法识别被投资企业名称列（请确认导出为对外投资表）")
        out = []
        for r in rows_iter:
            if not r:
                continue
            to_name = str(r[to_col]).strip() if to_col < len(r) and r[to_col] is not None else ""
            if not to_name:
                continue
            pct_val = r[pct_col] if (pct_col is not None and pct_col < len(r)) else None
            pct = _parse_float(pct_val)
            pct_text = str(pct_val).strip() if pct_val is not None else None
            out.append(_TyEdge(from_name=company_name, to_name=to_name, hold_pct=pct, hold_pct_text=pct_text))
        return out

    raise HTTPException(status_code=400, detail="kind must be shareholders|investments")


@router.post("/admin/import/targets-csv")
async def admin_import_targets_csv(
    snapshot_name: str = Query(..., min_length=1),
    file: UploadFile = File(...),
) -> dict[str, Any]:
    """
    方案B：上传目标公司清单（CSV）并导入为 targets + entities（最小字段）。

    CSV 表头建议：
    - name（必填）
    - credit_code（建议）
    - province / city（可选）
    - alias（可选）
    - is_key（可选：true/false/1/0）
    - notes（可选）
    """
    raw = await file.read()
    try:
        text_data = raw.decode("utf-8-sig")
    except Exception:
        text_data = raw.decode("utf-8", errors="replace")

    reader = csv.DictReader(io.StringIO(text_data))
    rows = list(reader)
    if not rows:
        raise HTTPException(status_code=400, detail="empty csv")

    inserted_entities = 0
    inserted_targets = 0

    with get_db() as db:
        db.execute(
            text(
                """
                INSERT INTO equity_snapshots (snapshot_name)
                VALUES (:s)
                ON CONFLICT (snapshot_name) DO NOTHING
                """
            ),
            {"s": snapshot_name},
        )

        for r in rows:
            name = _csv_get(r, "name")
            if not name:
                continue
            credit_code = _csv_get(r, "credit_code") or None
            province = _csv_get(r, "province") or None
            city = _csv_get(r, "city") or None
            alias = _csv_get(r, "alias") or None
            notes = _csv_get(r, "notes") or None
            is_key_raw = _csv_get(r, "is_key").lower()
            is_key = is_key_raw in {"1", "true", "yes", "y", "是"}

            entity_id = str(uuid.uuid4())
            db.execute(
                text(
                    """
                    INSERT INTO equity_entities
                      (id, snapshot_name, name, entity_type, credit_code, province, city, raw_source_json)
                    VALUES
                      (:id, :s, :name, 'company', :cc, :p, :c, :raw)
                    ON CONFLICT (id) DO NOTHING
                    """
                ),
                {
                    "id": entity_id,
                    "s": snapshot_name,
                    "name": name,
                    "cc": credit_code,
                    "p": province,
                    "c": city,
                    "raw": json.dumps({"import": "targets-csv", "alias": alias, "notes": notes}, ensure_ascii=False),
                },
            )
            inserted_entities += 1

            db.execute(
                text(
                    """
                    INSERT INTO equity_targets
                      (id, snapshot_name, entity_id, name, credit_code, alias, is_key, notes)
                    VALUES
                      (:id, :s, :eid, :name, :cc, :alias, :key, :notes)
                    ON CONFLICT (id) DO NOTHING
                    """
                ),
                {
                    "id": str(uuid.uuid4()),
                    "s": snapshot_name,
                    "eid": entity_id,
                    "name": name,
                    "cc": credit_code,
                    "alias": alias,
                    "key": is_key,
                    "notes": notes,
                },
            )
            inserted_targets += 1

    return {
        "ok": True,
        "snapshot_name": snapshot_name,
        "inserted": {"entities": inserted_entities, "targets": inserted_targets},
    }


@router.get("/snapshots/latest")
def latest_snapshot() -> dict[str, Any]:
    """
    返回最近创建的 snapshot_name（用于前端默认选中最新批次）。
    """
    with get_db() as db:
        row = db.execute(
            text(
                """
                SELECT snapshot_name
                FROM equity_snapshots
                ORDER BY created_at DESC NULLS LAST
                LIMIT 1
                """
            )
        ).fetchone()
    return {"snapshot_name": (str(row[0]) if row and row[0] else "")}


@router.post("/admin/import/edges-csv")
async def admin_import_edges_csv(
    snapshot_name: str = Query(..., min_length=1),
    file: UploadFile = File(...),
) -> dict[str, Any]:
    """
    上传股权边（CSV）并导入为 equity_edges；在当前 snapshot 内按名称匹配主体，不存在则自动创建主体。

    CSV 表头建议（最小）：
    - from_name（必填）
    - to_name（必填）
    - hold_pct（可选，0~100）
    - hold_pct_text（可选）
    - source_platform/source_doc/collected_at（可选）
    """
    raw = await file.read()
    try:
        text_data = raw.decode("utf-8-sig")
    except Exception:
        text_data = raw.decode("utf-8", errors="replace")

    reader = csv.DictReader(io.StringIO(text_data))
    rows = list(reader)
    if not rows:
        raise HTTPException(status_code=400, detail="empty csv")

    inserted_edges = 0
    created_entities = 0

    with get_db() as db:
        db.execute(
            text(
                """
                INSERT INTO equity_snapshots (snapshot_name)
                VALUES (:s)
                ON CONFLICT (snapshot_name) DO NOTHING
                """
            ),
            {"s": snapshot_name},
        )

        # 缓存：name -> entity_id（只在当前导入批次内使用）
        name_to_id: dict[str, str] = {}

        def _get_or_create_entity_id(name: str) -> str:
            nonlocal created_entities
            k = name.strip()
            if not k:
                return ""
            if k in name_to_id:
                return name_to_id[k]

            # 先查库（同 snapshot 同 name）
            r = db.execute(
                text(
                    """
                    SELECT id
                    FROM equity_entities
                    WHERE snapshot_name=:s AND name=:n
                    ORDER BY created_at ASC
                    LIMIT 1
                    """
                ),
                {"s": snapshot_name, "n": k},
            ).fetchone()
            if r and r[0]:
                eid = str(r[0])
                name_to_id[k] = eid
                return eid

            # 不存在则创建（最小字段）
            eid = str(uuid.uuid4())
            db.execute(
                text(
                    """
                    INSERT INTO equity_entities
                      (id, snapshot_name, name, entity_type, raw_source_json)
                    VALUES
                      (:id, :s, :name, 'company', :raw)
                    ON CONFLICT (id) DO NOTHING
                    """
                ),
                {"id": eid, "s": snapshot_name, "name": k, "raw": json.dumps({"import": "edges-csv"}, ensure_ascii=False)},
            )
            created_entities += 1
            name_to_id[k] = eid
            return eid

        for r in rows:
            from_name = _csv_get(r, "from_name")
            to_name = _csv_get(r, "to_name")
            if not from_name or not to_name:
                continue

            from_id = _get_or_create_entity_id(from_name)
            to_id = _get_or_create_entity_id(to_name)
            if not from_id or not to_id:
                continue

            hold_pct_raw = _csv_get(r, "hold_pct")
            hold_pct = None
            if hold_pct_raw:
                try:
                    hold_pct = float(hold_pct_raw)
                except Exception:
                    hold_pct = None
            hold_pct_text = _csv_get(r, "hold_pct_text") or None
            source_platform = _csv_get(r, "source_platform") or None
            source_doc = _csv_get(r, "source_doc") or None
            collected_at = _csv_get(r, "collected_at") or None

            db.execute(
                text(
                    """
                    INSERT INTO equity_edges
                      (id, snapshot_name, from_entity_id, to_entity_id, hold_pct, hold_pct_text,
                       source_platform, source_doc, collected_at)
                    VALUES
                      (:id, :s, :f, :t, :hp, :hpt, :sp, :sd, :ca)
                    ON CONFLICT (id) DO NOTHING
                    """
                ),
                {
                    "id": str(uuid.uuid4()),
                    "s": snapshot_name,
                    "f": from_id,
                    "t": to_id,
                    "hp": hold_pct,
                    "hpt": hold_pct_text,
                    "sp": source_platform,
                    "sd": source_doc,
                    "ca": collected_at,
                },
            )
            inserted_edges += 1

    return {"ok": True, "snapshot_name": snapshot_name, "inserted": {"edges": inserted_edges, "created_entities": created_entities}}


@router.post("/admin/import/bundle-zip")
async def admin_import_bundle_zip(
    snapshot_name: str = Query(..., min_length=1),
    file: UploadFile = File(...),
) -> dict[str, Any]:
    """
    上传“资料包 zip”一键导入（本地汇析产物包）。

    约定：zip 内至少包含以下文件中的一个或多个（文件可在任意子目录，按文件名匹配）：
    - targets.csv（可选）
    - entities.csv（可选，但没有 entities 将无法导入 edges/targets 的 entity_id）
    - equity_edges.csv（可选）

    CSV 字段建议：
    - targets.csv: name, credit_code?, alias?, is_key?, notes?, province?, city?（后两者会写入 entity）
    - entities.csv: id?, name, entity_type?, credit_code?, province?, city?, district?, reg_location?, industry?, status?, is_listed?, is_state_owned?, is_overseas?
    - equity_edges.csv: from_name?, to_name?, from_entity_id?, to_entity_id?, hold_pct?, hold_pct_text?
      - 优先使用 *_entity_id；若缺失则尝试用 name 在当前 snapshot 内匹配主体
    """
    raw = await file.read()
    _dbg(
        "Z0",
        "backend/routers/equity.py:admin_import_bundle_zip",
        "bundle upload received",
        {"snapshot_name": snapshot_name, "filename": file.filename, "size_bytes": len(raw)},
    )
    try:
        zf = zipfile.ZipFile(BytesIO(raw))
    except Exception as e:
        _dbg(
            "Z0",
            "backend/routers/equity.py:admin_import_bundle_zip",
            "invalid zip",
            {"err": str(e), "filename": file.filename, "size_bytes": len(raw)},
        )
        raise HTTPException(status_code=400, detail=f"invalid zip: {e}")

    infos = [i.filename for i in zf.infolist() if not i.is_dir()]
    _dbg(
        "Z1",
        "backend/routers/equity.py:admin_import_bundle_zip",
        "zip file list",
        {"count": len(infos), "files": infos[:50]},
    )

    targets_text = _read_zip_member_text(zf, "targets.csv")
    entities_text = _read_zip_member_text(zf, "entities.csv")
    edges_text = _read_zip_member_text(zf, "equity_edges.csv")

    _dbg(
        "Z1",
        "backend/routers/equity.py:admin_import_bundle_zip",
        "zip members matched",
        {
            "has_targets": bool(targets_text),
            "has_entities": bool(entities_text),
            "has_edges": bool(edges_text),
            "targets_head": (targets_text.splitlines()[:2] if targets_text else None),
            "entities_head": (entities_text.splitlines()[:2] if entities_text else None),
            "edges_head": (edges_text.splitlines()[:2] if edges_text else None),
        },
    )

    if not any([targets_text, entities_text, edges_text]):
        raise HTTPException(status_code=400, detail="zip missing targets.csv/entities.csv/equity_edges.csv")

    inserted_entities = 0
    inserted_targets = 0
    inserted_edges = 0
    created_entities = 0

    try:
        with get_db() as db:
            _dbg(
                "Z1",
                "backend/routers/equity.py:admin_import_bundle_zip",
                "db import start",
                {"snapshot_name": snapshot_name},
            )
            db.execute(
                text(
                    """
                    INSERT INTO equity_snapshots (snapshot_name)
                    VALUES (:s)
                    ON CONFLICT (snapshot_name) DO NOTHING
                    """
                ),
                {"s": snapshot_name},
            )

            # name -> entity_id cache
            name_to_id: dict[str, str] = {}

            def _get_or_create_entity_id(name: str) -> str:
                nonlocal created_entities
                k = (name or "").strip()
                if not k:
                    return ""
                if k in name_to_id:
                    return name_to_id[k]
                r = db.execute(
                    text(
                        """
                        SELECT id
                        FROM equity_entities
                        WHERE snapshot_name=:s AND name=:n
                        ORDER BY created_at ASC
                        LIMIT 1
                        """
                    ),
                    {"s": snapshot_name, "n": k},
                ).fetchone()
                if r and r[0]:
                    eid = str(r[0])
                    name_to_id[k] = eid
                    return eid
                eid = str(uuid.uuid4())
                db.execute(
                    text(
                        """
                        INSERT INTO equity_entities
                          (id, snapshot_name, name, entity_type, raw_source_json)
                        VALUES
                          (:id, :s, :name, 'company', :raw)
                        ON CONFLICT (id) DO NOTHING
                        """
                    ),
                    {"id": eid, "s": snapshot_name, "name": k, "raw": json.dumps({"import": "bundle-zip"}, ensure_ascii=False)},
                )
                created_entities += 1
                name_to_id[k] = eid
                return eid

            # 1) entities.csv
            if entities_text:
                rdr = csv.DictReader(io.StringIO(entities_text))
                for r in rdr:
                    name = _csv_get(r, "name")
                    if not name:
                        continue
                    eid = _csv_get(r, "id") or str(uuid.uuid4())
                    entity_type = _csv_get(r, "entity_type") or "company"
                    credit_code = _csv_get(r, "credit_code") or None
                    province = _csv_get(r, "province") or None
                    city = _csv_get(r, "city") or None
                    district = _csv_get(r, "district") or None
                    reg_location = _csv_get(r, "reg_location") or None
                    industry = _csv_get(r, "industry") or None
                    status = _csv_get(r, "status") or None
                    raw_in = _csv_get(r, "raw_source_json") or _csv_get(r, "profile_json") or ""
                    is_listed = _csv_get(r, "is_listed").lower()
                    is_state_owned = _csv_get(r, "is_state_owned").lower()
                    is_overseas = _csv_get(r, "is_overseas").lower()

                    if reg_location and (not city or not province or not district):
                        p2, c2, d2 = normalize_geo(
                            reg_location,
                            city or "",
                            district or "",
                            hint_province=province,
                        )
                        if not province and p2:
                            province = p2
                        if not city and c2:
                            city = c2
                        if not district and d2:
                            district = d2

                    def _to_bool(x: str) -> bool | None:
                        if not x:
                            return None
                        if x in {"1", "true", "yes", "y", "是"}:
                            return True
                        if x in {"0", "false", "no", "n", "否"}:
                            return False
                        return None

                    db.execute(
                        text(
                            """
                            INSERT INTO equity_entities
                              (id, snapshot_name, name, entity_type, credit_code, province, city, district, reg_location,
                               industry, status, is_listed, is_state_owned, is_overseas, raw_source_json)
                            VALUES
                              (:id, :s, :name, :t, :cc, :p, :c, :d, :rl, :ind, :st, :l, :so, :ov, :raw)
                            ON CONFLICT (id) DO UPDATE
                            SET snapshot_name=EXCLUDED.snapshot_name,
                                name=EXCLUDED.name,
                                entity_type=EXCLUDED.entity_type,
                                credit_code=EXCLUDED.credit_code,
                                province=EXCLUDED.province,
                                city=EXCLUDED.city,
                                district=EXCLUDED.district,
                                reg_location=EXCLUDED.reg_location,
                                industry=EXCLUDED.industry,
                                status=EXCLUDED.status,
                                is_listed=EXCLUDED.is_listed,
                                is_state_owned=EXCLUDED.is_state_owned,
                                is_overseas=EXCLUDED.is_overseas,
                                raw_source_json=EXCLUDED.raw_source_json
                            """
                        ),
                        {
                            "id": eid,
                            "s": snapshot_name,
                            "name": name,
                            "t": entity_type,
                            "cc": credit_code,
                            "p": province,
                            "c": city,
                            "d": district,
                            "rl": reg_location,
                            "ind": industry,
                            "st": status,
                            "l": _to_bool(is_listed),
                            "so": _to_bool(is_state_owned),
                            "ov": _to_bool(is_overseas),
                            "raw": raw_in if raw_in else json.dumps({"import": "bundle-zip"}, ensure_ascii=False),
                        },
                    )
                    inserted_entities += 1
                    name_to_id[name] = eid

            # 2) targets.csv
            if targets_text:
                rdr = csv.DictReader(io.StringIO(targets_text))
                for r in rdr:
                    name = _csv_get(r, "name")
                    if not name:
                        continue
                    credit_code = _csv_get(r, "credit_code") or None
                    alias = _csv_get(r, "alias") or None
                    notes = _csv_get(r, "notes") or None
                    is_key_raw = _csv_get(r, "is_key").lower()
                    is_key = is_key_raw in {"1", "true", "yes", "y", "是"}
                    province = _csv_get(r, "province") or None
                    city = _csv_get(r, "city") or None

                    eid = _get_or_create_entity_id(name)
                    # 尝试把 province/city 回填到 entity（若 entity 缺失）
                    if province or city:
                        db.execute(
                            text(
                                """
                                UPDATE equity_entities
                                SET province = COALESCE(province, :p),
                                    city = COALESCE(city, :c)
                                WHERE snapshot_name=:s AND id=:id
                                """
                            ),
                            {"s": snapshot_name, "id": eid, "p": province, "c": city},
                        )

                    _upsert_equity_target(
                        db,
                        snapshot_name=snapshot_name,
                        entity_id=eid,
                        name=name,
                        credit_code=credit_code,
                        alias=alias,
                        is_key=is_key,
                        notes=notes,
                    )
                    inserted_targets += 1

            # 3) equity_edges.csv
            if edges_text:
                rdr = csv.DictReader(io.StringIO(edges_text))
                for r in rdr:
                    from_eid = _csv_get(r, "from_entity_id")
                    to_eid = _csv_get(r, "to_entity_id")
                    if not from_eid:
                        from_name = _csv_get(r, "from_name")
                        from_eid = _get_or_create_entity_id(from_name)
                    if not to_eid:
                        to_name = _csv_get(r, "to_name")
                        to_eid = _get_or_create_entity_id(to_name)
                    if not from_eid or not to_eid:
                        continue

                    hp = _parse_float(_csv_get(r, "hold_pct"))
                    hpt = _csv_get(r, "hold_pct_text") or None
                    db.execute(
                        text(
                            """
                            INSERT INTO equity_edges
                              (id, snapshot_name, from_entity_id, to_entity_id, hold_pct, hold_pct_text, source_platform)
                            VALUES
                              (:id, :s, :f, :t, :hp, :hpt, :src)
                            ON CONFLICT (id) DO NOTHING
                            """
                        ),
                        {
                            "id": str(uuid.uuid4()),
                            "s": snapshot_name,
                            "f": from_eid,
                            "t": to_eid,
                            "hp": hp,
                            "hpt": hpt,
                            "src": "bundle",
                        },
                    )
                    inserted_edges += 1

            _dbg(
                "Z1",
                "backend/routers/equity.py:admin_import_bundle_zip",
                "db import done",
                {
                    "snapshot_name": snapshot_name,
                    "inserted": {
                        "entities": inserted_entities,
                        "targets": inserted_targets,
                        "edges": inserted_edges,
                        "created_entities": created_entities,
                    },
                },
            )
    except HTTPException:
        raise
    except Exception as e:
        _dbg(
            "Z2",
            "backend/routers/equity.py:admin_import_bundle_zip",
            "unhandled exception during import",
            {
                "err": repr(e),
                "snapshot_name": snapshot_name,
                "has_targets": bool(targets_text),
                "has_entities": bool(entities_text),
                "has_edges": bool(edges_text),
            },
        )
        raise HTTPException(status_code=500, detail="internal error while importing bundle-zip")

    return {
        "ok": True,
        "snapshot_name": snapshot_name,
        "inserted": {
            "entities": inserted_entities,
            "targets": inserted_targets,
            "edges": inserted_edges,
            "created_entities": created_entities,
        },
    }


@router.post("/admin/import/tianyancha-xlsx")
async def admin_import_tianyancha_xlsx(
    snapshot_name: str = Query(..., min_length=1),
    company_name: str = Query(..., min_length=1),
    shareholders_xlsx: UploadFile | None = File(None),
    investments_xlsx: UploadFile | None = File(None),
) -> dict[str, Any]:
    """
    天眼查半自动：上传“股东信息.xlsx”和“对外投资.xlsx”，自动产出 entities + edges，并确保 company_name 在 targets 中存在。
    """
    comp = company_name.strip()
    if not comp:
        raise HTTPException(status_code=400, detail="company_name required")
    if shareholders_xlsx is None and investments_xlsx is None:
        raise HTTPException(status_code=400, detail="need at least one xlsx file")

    edges: list[_TyEdge] = []
    if shareholders_xlsx is not None:
        raw = await shareholders_xlsx.read()
        edges.extend(_read_tianyancha_xlsx_edges(raw, comp, kind="shareholders"))
    if investments_xlsx is not None:
        raw = await investments_xlsx.read()
        edges.extend(_read_tianyancha_xlsx_edges(raw, comp, kind="investments"))

    if not edges:
        raise HTTPException(status_code=400, detail="no edges parsed from xlsx")

    inserted_edges = 0
    created_entities = 0

    with get_db() as db:
        db.execute(
            text(
                """
                INSERT INTO equity_snapshots (snapshot_name)
                VALUES (:s)
                ON CONFLICT (snapshot_name) DO NOTHING
                """
            ),
            {"s": snapshot_name},
        )

        name_to_id: dict[str, str] = {}

        def _get_or_create_entity_id(name: str) -> str:
            nonlocal created_entities
            k = name.strip()
            if not k:
                return ""
            if k in name_to_id:
                return name_to_id[k]
            r = db.execute(
                text(
                    """
                    SELECT id
                    FROM equity_entities
                    WHERE snapshot_name=:s AND name=:n
                    ORDER BY created_at ASC
                    LIMIT 1
                    """
                ),
                {"s": snapshot_name, "n": k},
            ).fetchone()
            if r and r[0]:
                eid = str(r[0])
                name_to_id[k] = eid
                return eid
            eid = str(uuid.uuid4())
            db.execute(
                text(
                    """
                    INSERT INTO equity_entities
                      (id, snapshot_name, name, entity_type, raw_source_json)
                    VALUES
                      (:id, :s, :name, 'company', :raw)
                    ON CONFLICT (id) DO NOTHING
                    """
                ),
                {"id": eid, "s": snapshot_name, "name": k, "raw": json.dumps({"import": "tianyancha-xlsx"}, ensure_ascii=False)},
            )
            created_entities += 1
            name_to_id[k] = eid
            return eid

        # 确保公司自身存在 & 在 targets 中存在（若无则补一条 target）
        company_id = _get_or_create_entity_id(comp)
        # targets 里可能已经有（来自 targets-csv）；若无则插入
        exists = db.execute(
            text(
                """
                SELECT 1
                FROM equity_targets
                WHERE snapshot_name=:s AND name=:n
                LIMIT 1
                """
            ),
            {"s": snapshot_name, "n": comp},
        ).fetchone()
        if not exists:
            db.execute(
                text(
                    """
                    INSERT INTO equity_targets
                      (id, snapshot_name, entity_id, name, is_key, notes)
                    VALUES
                      (:id, :s, :eid, :name, FALSE, :notes)
                    ON CONFLICT (id) DO NOTHING
                    """
                ),
                {"id": str(uuid.uuid4()), "s": snapshot_name, "eid": company_id, "name": comp, "notes": "tianyancha-xlsx auto target"},
            )

        for e in edges:
            from_id = _get_or_create_entity_id(e.from_name)
            to_id = _get_or_create_entity_id(e.to_name)
            if not from_id or not to_id:
                continue
            db.execute(
                text(
                    """
                    INSERT INTO equity_edges
                      (id, snapshot_name, from_entity_id, to_entity_id, hold_pct, hold_pct_text,
                       source_platform)
                    VALUES
                      (:id, :s, :f, :t, :hp, :hpt, 'tianyancha')
                    ON CONFLICT (id) DO NOTHING
                    """
                ),
                {
                    "id": str(uuid.uuid4()),
                    "s": snapshot_name,
                    "f": from_id,
                    "t": to_id,
                    "hp": e.hold_pct,
                    "hpt": e.hold_pct_text,
                },
            )
            inserted_edges += 1

    return {
        "ok": True,
        "snapshot_name": snapshot_name,
        "company_name": comp,
        "parsed_edges": len(edges),
        "inserted": {"edges": inserted_edges, "created_entities": created_entities},
    }


@router.get("/targets")
def targets(snapshot_name: str = Query(..., min_length=1)) -> dict[str, Any]:
    with get_db() as db:
        rows = db.execute(
            text(
                """
                SELECT DISTINCT ON (entity_id)
                  id, entity_id, name, credit_code, alias, is_key
                FROM equity_targets
                WHERE snapshot_name = :s
                ORDER BY entity_id, is_key DESC, name ASC
                """
            ),
            {"s": snapshot_name},
        ).fetchall()
    return {
        "snapshot": {"name": snapshot_name},
        "items": [
            {
                "id": str(r[0]),
                "entity_id": str(r[1]),
                "name": str(r[2]),
                "credit_code": r[3],
                "alias": r[4],
                "is_key": bool(r[5]),
            }
            for r in rows
        ],
    }


def _graph_geo_for_api(province: Any, city: Any, reg_location: Any, name: str) -> dict[str, Any]:
    """图谱/地图 API：返回展示用省、市（含注册地推断 + 公司名弱提示）。"""
    nm = str(name or "")
    rp = str(province or "").strip()
    rc = str(city or "").strip()
    rr_raw = reg_location if reg_location is not None else ""
    rr = str(rr_raw).strip()
    dp = province_display_label(rp or None, rr or None, company_name=nm)
    dc = city_display_label(rc or None, rr or None, rp or None, company_name=nm)
    return {"province": dp, "city": dc, "reg_location": rr_raw}


def _apply_display_geo_to_graph_nodes(nodes: list[dict[str, Any]]) -> None:
    for n in nodes:
        g = n.get("geo") or {}
        n["geo"] = _graph_geo_for_api(g.get("province"), g.get("city"), g.get("reg_location"), str(n.get("name") or ""))


def _snapshot_target_entity_ids(snapshot_name: str) -> set[str]:
    with get_db() as db:
        rows = db.execute(
            text("SELECT entity_id FROM equity_targets WHERE snapshot_name = :s"),
            {"s": snapshot_name},
        ).fetchall()
    return {str(r[0]) for r in rows}


def _merge_two_graph_nodes(a: dict[str, Any], b: dict[str, Any]) -> dict[str, Any]:
    out = dict(a)
    if len(str(b.get("name") or "")) > len(str(out.get("name") or "")):
        out["name"] = b.get("name")
    cc_a = str(out.get("credit_code") or "").strip()
    cc_b = str(b.get("credit_code") or "").strip()
    out["credit_code"] = cc_a or cc_b or out.get("credit_code")
    out["entity_type"] = out.get("entity_type") or b.get("entity_type")
    g0 = dict(out.get("geo") or {})
    g1 = dict(b.get("geo") or {})
    for k in ("province", "city", "reg_location"):
        v0 = str(g0.get(k) or "").strip()
        v1 = str(g1.get(k) or "").strip()
        if not v0 and v1:
            g0[k] = g1[k]
    out["geo"] = g0
    if not str(out.get("industry") or "").strip() and str(b.get("industry") or "").strip():
        out["industry"] = b.get("industry")
    t0 = dict(out.get("tags") or {})
    t1 = dict(b.get("tags") or {})
    t0["is_target"] = bool(t0.get("is_target")) or bool(t1.get("is_target"))
    t0["is_key_target"] = bool(t0.get("is_key_target")) or bool(t1.get("is_key_target"))
    out["tags"] = t0
    return out


def _dedupe_graph_by_credit_code(
    nodes: list[dict[str, Any]],
    edges: list[dict[str, Any]],
    *,
    preferred_entity_id: str,
    target_entity_ids: set[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """
    同一统一社会信用代码在图谱里可能对应多个天眼 id；合并为单一节点并重写边，避免全景图重复。
    """
    by_cc: dict[str, list[dict[str, Any]]] = defaultdict(list)
    no_cc: list[dict[str, Any]] = []
    for n in nodes:
        cc = str(n.get("credit_code") or "").strip()
        if not cc:
            no_cc.append(n)
        else:
            by_cc[cc].append(n)

    id_remap: dict[str, str] = {}
    merged: dict[str, dict[str, Any]] = {}

    for n in no_cc:
        nid = str(n["id"])
        id_remap[nid] = nid
        merged[nid] = n

    for _cc, group in by_cc.items():
        ids = [str(x["id"]) for x in group]
        if len(group) == 1:
            nid = ids[0]
            id_remap[nid] = nid
            merged[nid] = group[0]
            continue
        canon = preferred_entity_id if preferred_entity_id in ids else ""
        if not canon:
            for tid in target_entity_ids:
                if tid in ids:
                    canon = tid
                    break
        if not canon:
            canon = min(ids)
        base = next(x for x in group if str(x["id"]) == canon)
        acc = dict(base)
        for x in group:
            if str(x["id"]) == canon:
                continue
            acc = _merge_two_graph_nodes(acc, x)
        acc["id"] = canon
        for oid in ids:
            id_remap[oid] = canon
        merged[canon] = acc

    new_edges: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for e in edges:
        fid = id_remap.get(str(e["from"]), str(e["from"]))
        tid = id_remap.get(str(e["to"]), str(e["to"]))
        if fid == tid:
            continue
        hp = e.get("hold_pct_text")
        hpk = str(hp) if hp is not None else ""
        hpn = str(e.get("hold_pct") if e.get("hold_pct") is not None else "")
        key = (fid, tid, hpk, hpn)
        if key in seen:
            continue
        seen.add(key)
        ne = dict(e)
        ne["from"] = fid
        ne["to"] = tid
        new_edges.append(ne)
    return list(merged.values()), new_edges


@router.get("/targets/geo")
def targets_geo(snapshot_name: str = Query(..., min_length=1)) -> dict[str, Any]:
    """
    目标公司清单（含地理信息），用于前端地图/聚合展示。

    返回每个 target 的 entity_id + province/city（来自 equity_entities）。
    """
    with get_db() as db:
        rows = db.execute(
            text(
                """
                SELECT DISTINCT ON (t.entity_id)
                  t.id, t.entity_id, t.name, t.credit_code, t.alias, t.is_key,
                  e.province, e.city, e.reg_location
                FROM equity_targets t
                JOIN equity_entities e
                  ON e.snapshot_name = t.snapshot_name
                 AND e.id = t.entity_id
                WHERE t.snapshot_name = :s
                ORDER BY t.entity_id, t.is_key DESC, t.name ASC
                """
            ),
            {"s": snapshot_name},
        ).fetchall()
    return {
        "snapshot": {"name": snapshot_name},
        "items": [
            {
                "id": str(r[0]),
                "entity_id": str(r[1]),
                "name": str(r[2]),
                "credit_code": r[3],
                "alias": r[4],
                "is_key": bool(r[5]),
                "geo": _graph_geo_for_api(r[6], r[7], r[8], str(r[2])),
            }
            for r in rows
        ],
    }


@router.get("/entities/geo")
def entities_geo(
    snapshot_name: str = Query(..., min_length=1),
    limit: int = Query(5000, ge=1, le=20000),
    include_unknown: bool = Query(True),
) -> dict[str, Any]:
    """
    全量主体（含地理信息），用于 equity 地图散点（每家公司一个点）。
    - 从 equity_entities 读取 province/city/reg_location（按市落点优先）
    - LEFT JOIN equity_targets 以标记 is_target/is_key_target
    """
    where_geo = (
        ""
        if include_unknown
        else (
            "AND (COALESCE(TRIM(e.province), '') <> '' OR COALESCE(TRIM(e.city), '') <> '' "
            "OR COALESCE(TRIM(e.reg_location), '') <> '')"
        )
    )
    sql = f"""
      SELECT
        e.id, e.name, e.entity_type, e.credit_code, e.province, e.city, e.reg_location,
        COALESCE(t.is_key, false) AS is_key_target,
        CASE WHEN t.entity_id IS NULL THEN false ELSE true END AS is_target
      FROM equity_entities e
      LEFT JOIN (
        SELECT DISTINCT ON (entity_id)
          entity_id, is_key
        FROM equity_targets
        WHERE snapshot_name = :s
        ORDER BY entity_id, is_key DESC
      ) t ON t.entity_id = e.id
      WHERE e.snapshot_name = :s
      {where_geo}
      ORDER BY
        COALESCE(t.is_key, false) DESC,
        e.name ASC
      LIMIT :lim
    """
    with get_db() as db:
        rows = db.execute(text(sql), {"s": snapshot_name, "lim": limit}).fetchall()
    return {
        "snapshot": {"name": snapshot_name},
        "items": [
            {
                "entity_id": str(r[0]),
                "name": str(r[1]),
                "entity_type": str(r[2]),
                "credit_code": r[3],
                "geo": _graph_geo_for_api(r[4], r[5], r[6], str(r[1])),
                "is_target": bool(r[8]),
                "is_key_target": bool(r[7]),
            }
            for r in rows
        ],
    }


@router.get("/entities/search")
def entities_search(
    snapshot_name: str = Query(..., min_length=1),
    q: str = Query(..., min_length=1),
    limit: int = Query(20, ge=1, le=50),
) -> dict[str, Any]:
    qq = f"%{q.strip()}%"
    with get_db() as db:
        rows = db.execute(
            text(
                """
                SELECT id, name, credit_code, entity_type, province, city
                FROM equity_entities
                WHERE snapshot_name = :s
                  AND (name ILIKE :q OR credit_code ILIKE :q)
                ORDER BY
                  CASE WHEN credit_code = :q_exact THEN 0 ELSE 1 END,
                  name ASC
                LIMIT :lim
                """
            ),
            {"s": snapshot_name, "q": qq, "q_exact": q.strip(), "lim": limit},
        ).fetchall()
    return {
        "items": [
            {
                "id": str(r[0]),
                "name": str(r[1]),
                "credit_code": r[2],
                "entity_type": str(r[3]),
                "province": r[4],
                "city": r[5],
            }
            for r in rows
        ]
    }


@router.get("/entities/{entity_id}")
def entity_detail(entity_id: str, snapshot_name: str = Query(..., min_length=1)) -> dict[str, Any]:
    with get_db() as db:
        row = db.execute(
            text(
                """
                SELECT id, name, credit_code, entity_type,
                       province, city, district, reg_location,
                       industry, status,
                       is_listed, is_state_owned, is_overseas
                FROM equity_entities
                WHERE snapshot_name = :s AND id = :id
                """
            ),
            {"s": snapshot_name, "id": entity_id},
        ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="entity not found")
    return {
        "id": str(row[0]),
        "name": str(row[1]),
        "credit_code": row[2],
        "entity_type": str(row[3]),
        "geo": {"province": row[4], "city": row[5], "district": row[6], "reg_location": row[7]},
        "industry": row[8],
        "status": row[9],
        "tags": {"is_listed": row[10], "is_state_owned": row[11], "is_overseas": row[12]},
    }


@router.get("/graph/ownership")
def graph_ownership(
    snapshot_name: str = Query(..., min_length=1),
    target_entity_id: str = Query(..., min_length=1),
    direction: str = Query("down", pattern="^(up|down)$"),
    min_pct: float = Query(0.2, ge=0.0, le=1.0),
    max_depth: int = Query(3, ge=1, le=10),
    max_nodes: int = Query(500, ge=50, le=5000),
) -> dict[str, Any]:
    """
    MVP：用 BFS 展开子图，返回 nodes/edges。
    - down：from -> to
    - up：反向沿 to <- from 展开
    - 下游递归阈值：仅对 hold_pct >= min_pct*100 的边继续展开（但仍会返回已纳入子图的边）
    """
    threshold_pct = float(min_pct) * 100.0
    nodes: dict[str, dict[str, Any]] = {}
    edges: list[dict[str, Any]] = []

    truncated = False
    truncate_reason: str | None = None
    depth_max = 0

    frontier = [target_entity_id]
    visited = {target_entity_id}

    def _load_entity(eid: str) -> None:
        if eid in nodes:
            return
        with get_db() as db:
            r = db.execute(
                text(
                    """
                    SELECT id, name, entity_type, credit_code, province, city, reg_location, industry,
                           is_listed, is_state_owned, is_overseas
                    FROM equity_entities
                    WHERE snapshot_name=:s AND id=:id
                    """
                ),
                {"s": snapshot_name, "id": eid},
            ).fetchone()
        if not r:
            return
        nodes[eid] = {
            "id": str(r[0]),
            "name": str(r[1]),
            "entity_type": str(r[2]),
            "credit_code": r[3],
            "tags": {"is_listed": r[8], "is_state_owned": r[9], "is_overseas": r[10]},
            "geo": {"province": r[4], "city": r[5], "reg_location": r[6]},
            "industry": r[7],
        }

    _load_entity(target_entity_id)

    for depth in range(1, max_depth + 1):
        if not frontier:
            break
        depth_max = depth
        next_frontier: list[str] = []
        with get_db() as db:
            if direction == "down":
                if len(frontier) == 0:
                    rel_rows = []
                else:
                    rel_rows = db.execute(
                        text(
                            """
                            SELECT e.id, e.from_entity_id, e.to_entity_id, e.hold_pct, e.hold_pct_text, e.source_platform
                            FROM equity_edges e
                            WHERE e.snapshot_name=:s AND e.from_entity_id IN :from_ids
                            """
                        ).bindparams(bindparam("from_ids", expanding=True)),
                        {"s": snapshot_name, "from_ids": list(frontier)},
                    ).fetchall()
            else:
                if len(frontier) == 0:
                    rel_rows = []
                else:
                    rel_rows = db.execute(
                        text(
                            """
                            SELECT e.id, e.from_entity_id, e.to_entity_id, e.hold_pct, e.hold_pct_text, e.source_platform
                            FROM equity_edges e
                            WHERE e.snapshot_name=:s AND e.to_entity_id IN :to_ids
                            """
                        ).bindparams(bindparam("to_ids", expanding=True)),
                        {"s": snapshot_name, "to_ids": list(frontier)},
                    ).fetchall()

        for r in rel_rows:
            edge_id = str(r[0])
            from_id = str(r[1])
            to_id = str(r[2])
            hold_pct = r[3]
            hold_pct_text = r[4]
            source = r[5]

            edges.append(
                {
                    "id": edge_id,
                    "from": from_id,
                    "to": to_id,
                    "hold_pct": hold_pct,
                    "hold_pct_text": hold_pct_text,
                    "source": source,
                }
            )

            # Determine neighbor directionally
            neighbor = to_id if direction == "down" else from_id

            # Heavy profile: unknown hold_pct 仍允许递归（否则很多边没有比例会导致“只有一家公司”）
            can_recurse = hold_pct is None or float(hold_pct) >= threshold_pct
            if not can_recurse:
                _load_entity(from_id)
                _load_entity(to_id)
                continue

            if neighbor not in visited:
                visited.add(neighbor)
                next_frontier.append(neighbor)
                _load_entity(from_id)
                _load_entity(to_id)

                if len(nodes) >= max_nodes:
                    truncated = True
                    truncate_reason = "max_nodes_exceeded"
                    break
        if truncated:
            break
        frontier = next_frontier

    targ_ids = _snapshot_target_entity_ids(snapshot_name)
    dn, de = _dedupe_graph_by_credit_code(
        list(nodes.values()),
        edges,
        preferred_entity_id=target_entity_id,
        target_entity_ids=targ_ids,
    )
    _apply_display_geo_to_graph_nodes(dn)

    return {
        "snapshot": {"name": snapshot_name},
        "params": {
            "direction": direction,
            "min_pct": min_pct,
            "max_depth": max_depth,
            "max_nodes": max_nodes,
        },
        "nodes": dn,
        "edges": de,
        "stats": {
            "node_count": len(dn),
            "edge_count": len(de),
            "depth_max": depth_max,
            "truncated": truncated,
            "truncate_reason": truncate_reason,
        },
    }


@router.get("/graph/panorama")
def graph_panorama(
    snapshot_name: str = Query(..., min_length=1),
    target_entity_id: str = Query(..., min_length=1),
    min_pct: float = Query(0.2, ge=0.0, le=1.0),
    max_depth: int = Query(4, ge=1, le=10),
    max_nodes: int = Query(1200, ge=50, le=10000),
) -> dict[str, Any]:
    """
    全景子图：上游 + 下游同时展开并合并（用于“从最上到最下”的架构图展示）。
    """
    down = graph_ownership(
        snapshot_name=snapshot_name,
        target_entity_id=target_entity_id,
        direction="down",
        min_pct=min_pct,
        max_depth=max_depth,
        max_nodes=max_nodes,
    )
    up = graph_ownership(
        snapshot_name=snapshot_name,
        target_entity_id=target_entity_id,
        direction="up",
        min_pct=min_pct,
        max_depth=max_depth,
        max_nodes=max_nodes,
    )

    with get_db() as db:
        trows = db.execute(
            text(
                """
                SELECT DISTINCT ON (entity_id) entity_id, is_key
                FROM equity_targets
                WHERE snapshot_name = :s
                ORDER BY entity_id, is_key DESC
                """
            ),
            {"s": snapshot_name},
        ).fetchall()
    target_map = {str(r[0]): bool(r[1]) for r in trows}

    nodes_by_id: dict[str, dict[str, Any]] = {}
    for n in (down.get("nodes") or []):
        nodes_by_id[str(n.get("id"))] = n
    for n in (up.get("nodes") or []):
        nid = str(n.get("id"))
        nodes_by_id[nid] = {**nodes_by_id.get(nid, {}), **n}

    for nid, n in nodes_by_id.items():
        tags = dict((n.get("tags") or {}))
        if nid in target_map:
            tags["is_target"] = True
            tags["is_key_target"] = bool(target_map[nid])
        n["tags"] = tags

    edges_by_id: dict[str, dict[str, Any]] = {}
    for e in (down.get("edges") or []):
        edges_by_id[str(e.get("id"))] = e
    for e in (up.get("edges") or []):
        edges_by_id[str(e.get("id"))] = e

    truncated = bool(down.get("stats", {}).get("truncated")) or bool(up.get("stats", {}).get("truncated"))
    truncate_reason = down.get("stats", {}).get("truncate_reason") or up.get("stats", {}).get("truncate_reason")
    depth_max = max(int(down.get("stats", {}).get("depth_max") or 0), int(up.get("stats", {}).get("depth_max") or 0))

    targ_ids = set(target_map.keys())
    dn, de = _dedupe_graph_by_credit_code(
        list(nodes_by_id.values()),
        list(edges_by_id.values()),
        preferred_entity_id=target_entity_id,
        target_entity_ids=targ_ids,
    )
    _apply_display_geo_to_graph_nodes(dn)

    return {
        "snapshot": {"name": snapshot_name},
        "params": {"min_pct": min_pct, "max_depth": max_depth, "max_nodes": max_nodes},
        "nodes": dn,
        "edges": de,
        "stats": {
            "node_count": len(dn),
            "edge_count": len(de),
            "depth_max": depth_max,
            "truncated": truncated,
            "truncate_reason": truncate_reason,
        },
    }


@router.get("/analysis/summary")
def analysis_summary(snapshot_name: str = Query(..., min_length=1)) -> dict[str, Any]:
    with get_db() as db:
        geo_rows = db.execute(
            text(
                """
                SELECT TRIM(COALESCE(city, '')), COALESCE(reg_location, ''), TRIM(COALESCE(province, '')), COALESCE(name, '')
                FROM equity_entities
                WHERE snapshot_name = :s
                """
            ),
            {"s": snapshot_name},
        ).fetchall()
        type_rows = db.execute(
            text(
                """
                SELECT entity_type AS k, COUNT(*) AS c
                FROM equity_entities
                WHERE snapshot_name = :s
                GROUP BY entity_type
                ORDER BY c DESC
                """
            ),
            {"s": snapshot_name},
        ).fetchall()

    city_ctr: Counter[str] = Counter()
    prov_ctr: Counter[str] = Counter()
    for cy, rl, pr, nm in geo_rows:
        cl = city_display_label(cy or None, rl or None, pr or None, company_name=str(nm or ""))
        city_ctr[cl if cl else "未知"] += 1
        pl = province_display_label(pr or None, rl or None, company_name=str(nm or ""))
        prov_ctr[pl if pl else "未知"] += 1

    city_sorted = dict(sorted(city_ctr.items(), key=lambda x: (-x[1], x[0])))
    prov_sorted = dict(sorted(prov_ctr.items(), key=lambda x: (-x[1], x[0])))

    return {
        "snapshot": {"name": snapshot_name},
        "distributions": {
            "city": city_sorted,
            "province": prov_sorted,
            "entity_type": {str(r[0]): int(r[1]) for r in type_rows},
        },
    }


@router.get("/analysis/compare")
def analysis_compare(
    snapshot_name: str = Query(..., min_length=1),
    entity_id_a: str = Query(..., min_length=1),
    entity_id_b: str = Query(..., min_length=1),
    min_pct: float = Query(0.2, ge=0.0, le=1.0),
    max_depth: int = Query(3, ge=1, le=10),
    max_nodes: int = Query(500, ge=50, le=5000),
) -> dict[str, Any]:
    """
    MVP 对比：用同一组参数展开 A/B 的子图，然后计算交集与差异统计。
    """
    ga = graph_ownership(
        snapshot_name=snapshot_name,
        target_entity_id=entity_id_a,
        direction="up",
        min_pct=min_pct,
        max_depth=max_depth,
        max_nodes=max_nodes,
    )
    gb = graph_ownership(
        snapshot_name=snapshot_name,
        target_entity_id=entity_id_b,
        direction="up",
        min_pct=min_pct,
        max_depth=max_depth,
        max_nodes=max_nodes,
    )

    nodes_a = {str(n.get("id")): n for n in (ga.get("nodes") or []) if isinstance(n, dict) and n.get("id")}
    nodes_b = {str(n.get("id")): n for n in (gb.get("nodes") or []) if isinstance(n, dict) and n.get("id")}

    common_ids = set(nodes_a.keys()).intersection(nodes_b.keys())
    common_shareholders = [
        {"entity_id": nid, "name": str((nodes_a.get(nid) or {}).get("name") or (nodes_b.get(nid) or {}).get("name") or nid)}
        for nid in list(common_ids)[:50]
        if nid not in {entity_id_a, entity_id_b}
    ]

    def _overseas_ratio(nodes: dict[str, dict[str, Any]]) -> float:
        if not nodes:
            return 0.0
        overseas = 0
        for n in nodes.values():
            tags = n.get("tags") or {}
            if isinstance(tags, dict) and tags.get("is_overseas") is True:
                overseas += 1
        return overseas / max(1, len(nodes))

    return {
        "snapshot": {"name": snapshot_name},
        "params": {"min_pct": min_pct, "max_depth": max_depth, "max_nodes": max_nodes},
        "a": {"entity_id": entity_id_a, "name": (nodes_a.get(entity_id_a) or {}).get("name") or entity_id_a},
        "b": {"entity_id": entity_id_b, "name": (nodes_b.get(entity_id_b) or {}).get("name") or entity_id_b},
        "overlap": {"common_shareholders": common_shareholders, "common_investments": []},
        "diff_stats": {
            "a_node_count": int((ga.get("stats") or {}).get("node_count") or 0),
            "b_node_count": int((gb.get("stats") or {}).get("node_count") or 0),
            "a_overseas_ratio": _overseas_ratio(nodes_a),
            "b_overseas_ratio": _overseas_ratio(nodes_b),
            "a_truncated": bool((ga.get("stats") or {}).get("truncated") or False),
            "b_truncated": bool((gb.get("stats") or {}).get("truncated") or False),
        },
    }


@router.get("/export/subgraph")
def export_subgraph(
    snapshot_name: str = Query(..., min_length=1),
    target_entity_id: str = Query(..., min_length=1),
    direction: str = Query("down", pattern="^(up|down)$"),
    min_pct: float = Query(0.2, ge=0.0, le=1.0),
    max_depth: int = Query(3, ge=1, le=10),
    max_nodes: int = Query(500, ge=50, le=5000),
) -> dict[str, Any]:
    payload = graph_ownership(
        snapshot_name=snapshot_name,
        target_entity_id=target_entity_id,
        direction=direction,
        min_pct=min_pct,
        max_depth=max_depth,
        max_nodes=max_nodes,
    )
    payload["exported_at"] = _now_iso()
    payload["export_id"] = str(uuid.uuid4())
    return payload


@router.get("/export/csv")
def export_csv(
    snapshot_name: str = Query(..., min_length=1),
    type: str = Query(..., pattern="^(entities|equity_edges|targets)$"),
) -> StreamingResponse:
    filename = f"{type}_{snapshot_name}.csv"
    with get_db() as db:
        if type == "entities":
            rows = db.execute(
                text(
                    """
                    SELECT snapshot_name, id, name, entity_type, credit_code, province, city, district, reg_location,
                           industry, status, established_date, is_listed, is_state_owned, is_overseas,
                           source_url, raw_source_json
                    FROM equity_entities
                    WHERE snapshot_name=:s
                    ORDER BY name ASC
                    """
                ),
                {"s": snapshot_name},
            ).fetchall()
            header = [
                "snapshot_name",
                "id",
                "name",
                "entity_type",
                "credit_code",
                "province",
                "city",
                "district",
                "reg_location",
                "industry",
                "status",
                "established_date",
                "is_listed",
                "is_state_owned",
                "is_overseas",
                "source_url",
                "raw_source_json",
            ]
        elif type == "targets":
            rows = db.execute(
                text(
                    """
                    SELECT DISTINCT ON (entity_id)
                      snapshot_name, id, entity_id, name, credit_code, alias, is_key, notes
                    FROM equity_targets
                    WHERE snapshot_name = :s
                    ORDER BY entity_id, is_key DESC, name ASC
                    """
                ),
                {"s": snapshot_name},
            ).fetchall()
            header = ["snapshot_name", "id", "entity_id", "name", "credit_code", "alias", "is_key", "notes"]
        else:
            rows = db.execute(
                text(
                    """
                    SELECT snapshot_name, id, from_entity_id, to_entity_id, hold_pct, hold_pct_text,
                           source_platform, source_doc, collected_at
                    FROM equity_edges
                    WHERE snapshot_name=:s
                    ORDER BY id ASC
                    """
                ),
                {"s": snapshot_name},
            ).fetchall()
            header = [
                "snapshot_name",
                "id",
                "from_entity_id",
                "to_entity_id",
                "hold_pct",
                "hold_pct_text",
                "source_platform",
                "source_doc",
                "collected_at",
            ]

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(header)
    for r in rows:
        w.writerow(list(r))
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/admin/import")
def admin_import(payload: dict[str, Any]) -> dict[str, Any]:
    """
    管理导入（MVP）：直接接收 JSON 结构写入 equity_* 表。

    请求格式（最小）：
    {
      "snapshot_name":"2026-04-08_run1",
      "entities":[{...}],
      "edges":[{...}],
      "targets":[{...}]  // 可选
    }
    """
    snapshot_name = str(payload.get("snapshot_name") or "").strip()
    if not snapshot_name:
        raise HTTPException(status_code=400, detail="snapshot_name required")
    entities = payload.get("entities") or []
    edges = payload.get("edges") or []
    targets = payload.get("targets") or []
    if not isinstance(entities, list) or not isinstance(edges, list) or not isinstance(targets, list):
        raise HTTPException(status_code=400, detail="entities/edges/targets must be lists")

    with get_db() as db:
        db.execute(
            text(
                """
                INSERT INTO equity_snapshots (snapshot_name)
                VALUES (:s)
                ON CONFLICT (snapshot_name) DO NOTHING
                """
            ),
            {"s": snapshot_name},
        )

        for e in entities:
            if not isinstance(e, dict):
                continue
            eid = str(e.get("id") or "").strip() or str(uuid.uuid4())
            p_ent = e.get("province")
            c_ent = e.get("city")
            d_ent = e.get("district")
            rl_ent = e.get("reg_location")
            if rl_ent and (not c_ent or not p_ent or not d_ent):
                p2, c2, d2 = normalize_geo(
                    str(rl_ent),
                    str(c_ent or ""),
                    str(d_ent or ""),
                    hint_province=p_ent,
                )
                if not p_ent and p2:
                    p_ent = p2
                if not c_ent and c2:
                    c_ent = c2
                if not d_ent and d2:
                    d_ent = d2
            db.execute(
                text(
                    """
                    INSERT INTO equity_entities
                      (id, snapshot_name, name, entity_type, credit_code, province, city, district, reg_location,
                       industry, status, established_date, is_listed, is_state_owned, is_overseas,
                       source_url, raw_source_json)
                    VALUES
                      (:id, :s, :name, :t, :cc, :p, :c, :d, :rl, :ind, :st, :ed, :l, :so, :ov, :url, :raw)
                    ON CONFLICT (id) DO UPDATE
                    SET snapshot_name=EXCLUDED.snapshot_name,
                        name=EXCLUDED.name,
                        entity_type=EXCLUDED.entity_type,
                        credit_code=EXCLUDED.credit_code,
                        province=EXCLUDED.province,
                        city=EXCLUDED.city,
                        district=EXCLUDED.district,
                        reg_location=EXCLUDED.reg_location,
                        industry=EXCLUDED.industry,
                        status=EXCLUDED.status,
                        established_date=EXCLUDED.established_date,
                        is_listed=EXCLUDED.is_listed,
                        is_state_owned=EXCLUDED.is_state_owned,
                        is_overseas=EXCLUDED.is_overseas,
                        source_url=EXCLUDED.source_url,
                        raw_source_json=EXCLUDED.raw_source_json
                    """
                ),
                {
                    "id": eid,
                    "s": snapshot_name,
                    "name": str(e.get("name") or "").strip(),
                    "t": str(e.get("entity_type") or "company").strip(),
                    "cc": e.get("credit_code"),
                    "p": p_ent,
                    "c": c_ent,
                    "d": d_ent,
                    "rl": rl_ent,
                    "ind": e.get("industry"),
                    "st": e.get("status"),
                    "ed": e.get("established_date"),
                    "l": e.get("is_listed"),
                    "so": e.get("is_state_owned"),
                    "ov": e.get("is_overseas"),
                    "url": e.get("source_url"),
                    "raw": json.dumps(e.get("raw_source") or {}, ensure_ascii=False),
                },
            )

        for ed in edges:
            if not isinstance(ed, dict):
                continue
            edge_id = str(ed.get("id") or "").strip() or str(uuid.uuid4())
            db.execute(
                text(
                    """
                    INSERT INTO equity_edges
                      (id, snapshot_name, from_entity_id, to_entity_id, hold_pct, hold_pct_text,
                       source_platform, source_doc, collected_at)
                    VALUES
                      (:id, :s, :f, :t, :hp, :hpt, :sp, :sd, :ca)
                    ON CONFLICT (id) DO UPDATE
                    SET snapshot_name=EXCLUDED.snapshot_name,
                        from_entity_id=EXCLUDED.from_entity_id,
                        to_entity_id=EXCLUDED.to_entity_id,
                        hold_pct=EXCLUDED.hold_pct,
                        hold_pct_text=EXCLUDED.hold_pct_text,
                        source_platform=EXCLUDED.source_platform,
                        source_doc=EXCLUDED.source_doc,
                        collected_at=EXCLUDED.collected_at
                    """
                ),
                {
                    "id": edge_id,
                    "s": snapshot_name,
                    "f": str(ed.get("from_entity_id") or "").strip(),
                    "t": str(ed.get("to_entity_id") or "").strip(),
                    "hp": ed.get("hold_pct"),
                    "hpt": ed.get("hold_pct_text"),
                    "sp": ed.get("source_platform"),
                    "sd": ed.get("source_doc"),
                    "ca": ed.get("collected_at"),
                },
            )

        for t in targets:
            if not isinstance(t, dict):
                continue
            eid_t = str(t.get("entity_id") or "").strip()
            if not eid_t:
                continue
            _upsert_equity_target(
                db,
                snapshot_name=snapshot_name,
                entity_id=eid_t,
                name=str(t.get("name") or "").strip(),
                credit_code=t.get("credit_code"),
                alias=t.get("alias"),
                is_key=bool(t.get("is_key") or False),
                notes=t.get("notes"),
            )

    return {"ok": True, "snapshot_name": snapshot_name, "entities": len(entities), "edges": len(edges), "targets": len(targets)}

